import os
import pandas as pd
import numpy as np
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tempfile
import io

analise_bp = Blueprint('analise', __name__)

UPLOAD_FOLDER = '/tmp/uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Criar pasta de upload se n√£o existir
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Fun√ß√µes de formata√ß√£o brasileira
def formatar_numero_brasileiro(valor):
    """
    Formata n√∫meros no padr√£o brasileiro (2.651,40)
    """
    if pd.isna(valor) or valor is None:
        return "0,00"
    
    # Converter para float se necess√°rio
    if isinstance(valor, (int, float, np.number)):
        numero = float(valor)
    else:
        try:
            numero = float(valor)
        except:
            return str(valor)
    
    # Formata√ß√£o brasileira
    if numero == int(numero):
        # N√∫mero inteiro
        return f"{int(numero):,}".replace(",", ".")
    else:
        # N√∫mero decimal
        return f"{numero:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_moeda_brasileira(valor):
    """
    Formata valores monet√°rios no padr√£o brasileiro (R$ 2.651,40)
    """
    numero_formatado = formatar_numero_brasileiro(valor)
    return f"R$ {numero_formatado}"

def formatar_percentual_brasileiro(valor):
    """
    Formata percentuais no padr√£o brasileiro (25,5%)
    """
    if pd.isna(valor) or valor is None:
        return "0,0%"
    
    numero = float(valor)
    return f"{numero:.1f}%".replace(".", ",")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def processar_arquivo_cargas(filepath):
    """
    Processa arquivo de cargas e gera an√°lise completa
    """
    try:
        # Carregar dados originais
        df_original = pd.read_excel(filepath, sheet_name='Agenda Recebimento')
        
        # Filtrar apenas "Em Aprova√ß√£o"
        df_aprovacao = df_original[df_original['Status'] == 'Em Aprova√ß√£o'].copy()
        
        if len(df_aprovacao) == 0:
            return None, "Nenhum registro encontrado com status 'Em Aprova√ß√£o'"
        
        # Converter dados num√©ricos
        df_aprovacao['Cobertura Atual'] = pd.to_numeric(df_aprovacao['Cobertura Atual'], errors='coerce')
        df_aprovacao['Saldo Pedido'] = pd.to_numeric(df_aprovacao['Saldo Pedido'], errors='coerce')
        df_aprovacao['Quantidade<br />Entrega'] = pd.to_numeric(df_aprovacao['Quantidade<br />Entrega'], errors='coerce')
        
        # Limpar dados
        df_clean = df_aprovacao[
            df_aprovacao['Cobertura Atual'].notna() & 
            df_aprovacao['Fornecedor'].notna() &
            df_aprovacao['Filial'].notna() &
            df_aprovacao['Mercadoria'].notna()
        ].copy()
        
        if len(df_clean) == 0:
            return None, "Nenhum registro v√°lido encontrado ap√≥s limpeza dos dados"
        
        # Gerar arquivo Excel
        output_file = gerar_excel_analise(df_clean)
        
        # Gerar resumo para resposta
        resumo = gerar_resumo_analise(df_clean)
        
        return output_file, resumo
        
    except Exception as e:
        return None, f"Erro ao processar arquivo: {str(e)}"

def gerar_excel_analise(df):
    """
    Gera arquivo Excel com an√°lise completa
    """
    
    # Criar workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # 1. ABA RESUMO EXECUTIVO
    ws_resumo = wb.create_sheet("üìä Resumo Executivo")
    
    # Calcular m√©tricas principais
    total_fornecedores = df['Fornecedor'].nunique()
    total_cargas = df['Carga'].nunique()
    total_itens = len(df)
    total_filiais = df['Filial'].nunique()
    valor_total = df['Saldo Pedido'].sum()
    cobertura_media_geral = df['Cobertura Atual'].mean()
    
    # An√°lise por faixas
    ate_44 = len(df[df['Cobertura Atual'] <= 44])
    entre_45_70 = len(df[(df['Cobertura Atual'] >= 45) & (df['Cobertura Atual'] <= 70)])
    acima_71 = len(df[df['Cobertura Atual'] >= 71])
    
    perc_ate_44 = (ate_44 / total_itens) * 100
    perc_45_70 = (entre_45_70 / total_itens) * 100
    perc_acima_71 = (acima_71 / total_itens) * 100
    
    # Dados do resumo com formata√ß√£o brasileira
    dados_resumo = [
        ["AN√ÅLISE DE CARGAS EM APROVA√á√ÉO", ""],
        [f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ""],
        ["", ""],
        ["M√âTRICAS GERAIS", ""],
        ["Total de Fornecedores", formatar_numero_brasileiro(total_fornecedores)],
        ["Total de Cargas", formatar_numero_brasileiro(total_cargas)],
        ["Total de Itens", formatar_numero_brasileiro(total_itens)],
        ["Total de Filiais", formatar_numero_brasileiro(total_filiais)],
        ["Valor Total", formatar_moeda_brasileira(valor_total)],
        ["Cobertura M√©dia Geral", f"{formatar_numero_brasileiro(cobertura_media_geral)} dias"],
        ["", ""],
        ["DISTRIBUI√á√ÉO POR FAIXAS DE COBERTURA", ""],
        ["At√© 44 dias", f"{formatar_numero_brasileiro(ate_44)} itens ({formatar_percentual_brasileiro(perc_ate_44)})"],
        ["Entre 45-70 dias", f"{formatar_numero_brasileiro(entre_45_70)} itens ({formatar_percentual_brasileiro(perc_45_70)})"],
        ["Acima de 71 dias", f"{formatar_numero_brasileiro(acima_71)} itens ({formatar_percentual_brasileiro(perc_acima_71)})"],
        ["", ""],
        ["DISTRIBUI√á√ÉO POR FILIAL", ""],
    ]
    
    # Adicionar dados por filial com formata√ß√£o brasileira
    for filial in df['Filial'].unique():
        dados_filial = df[df['Filial'] == filial]
        cobertura_filial = dados_filial['Cobertura Atual'].mean()
        valor_filial = dados_filial['Saldo Pedido'].sum()
        dados_resumo.append([filial, f"{formatar_numero_brasileiro(len(dados_filial))} itens, {formatar_numero_brasileiro(cobertura_filial)} dias, {formatar_moeda_brasileira(valor_filial)}"])
    
    # Preencher dados
    for row, (label, valor) in enumerate(dados_resumo, 1):
        cell_a = ws_resumo.cell(row=row, column=1, value=label)
        cell_b = ws_resumo.cell(row=row, column=2, value=valor)
        
        # Formata√ß√£o
        if "AN√ÅLISE DE CARGAS" in str(label):
            cell_a.font = Font(bold=True, size=16, color="FFFFFF")
            cell_a.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        elif label in ["M√âTRICAS GERAIS", "DISTRIBUI√á√ÉO POR FAIXAS DE COBERTURA", "DISTRIBUI√á√ÉO POR FILIAL"]:
            cell_a.font = Font(bold=True, size=12)
            cell_a.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_b.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar larguras
    ws_resumo.column_dimensions['A'].width = 40
    ws_resumo.column_dimensions['B'].width = 30
    
    # 2. ABA AN√ÅLISE POR FORNECEDOR
    ws_fornecedores = wb.create_sheet("üè≠ An√°lise por Fornecedor")
    
    # Cabe√ßalhos
    headers = [
        'Fornecedor', 'Total Itens', 'Total Cargas', 'Filiais Atendidas', 
        'Cobertura M√©dia', 'Valor Total (R$)', '% At√© 44 dias', 
        '% Entre 45-70 dias', '% Acima 71 dias', 'Recomenda√ß√£o'
    ]
    
    # Criar cabe√ßalho
    for col, header in enumerate(headers, 1):
        cell = ws_fornecedores.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Analisar cada fornecedor
    fornecedores_analise = []
    
    for fornecedor in df['Fornecedor'].unique():
        dados_forn = df[df['Fornecedor'] == fornecedor]
        
        total_itens = len(dados_forn)
        total_cargas = dados_forn['Carga'].nunique()
        filiais = dados_forn['Filial'].nunique()
        cobertura_media = dados_forn['Cobertura Atual'].mean()
        valor_total = dados_forn['Saldo Pedido'].sum()
        
        # Distribui√ß√£o por faixas
        ate_44 = len(dados_forn[dados_forn['Cobertura Atual'] <= 44])
        entre_45_70 = len(dados_forn[(dados_forn['Cobertura Atual'] >= 45) & (dados_forn['Cobertura Atual'] <= 70)])
        acima_71 = len(dados_forn[dados_forn['Cobertura Atual'] >= 71])
        
        perc_ate_44 = (ate_44 / total_itens) * 100
        perc_45_70 = (entre_45_70 / total_itens) * 100
        perc_acima_71 = (acima_71 / total_itens) * 100
        
        # Recomenda√ß√£o
        if perc_acima_71 > 50 or cobertura_media > 100:
            recomendacao = "‚ùå REJEITAR"
        elif perc_acima_71 > 25 or cobertura_media > 70:
            recomendacao = "‚ö†Ô∏è REVISAR"
        else:
            recomendacao = "‚úÖ APROVAR"
        
        fornecedores_analise.append({
            'Fornecedor': fornecedor,
            'Total_Itens': total_itens,
            'Total_Cargas': total_cargas,
            'Filiais': filiais,
            'Cobertura_Media': cobertura_media,
            'Valor_Total': valor_total,
            'Perc_Ate_44': perc_ate_44,
            'Perc_45_70': perc_45_70,
            'Perc_Acima_71': perc_acima_71,
            'Recomendacao': recomendacao
        })
    
    # Ordenar por cobertura m√©dia (mais cr√≠ticos primeiro)
    fornecedores_analise.sort(key=lambda x: x['Cobertura_Media'], reverse=True)
    
    # Preencher dados com formata√ß√£o brasileira
    for row, forn in enumerate(fornecedores_analise, 2):
        dados = [
            forn['Fornecedor'][:40],
            formatar_numero_brasileiro(forn['Total_Itens']),
            formatar_numero_brasileiro(forn['Total_Cargas']),
            formatar_numero_brasileiro(forn['Filiais']),
            formatar_numero_brasileiro(forn['Cobertura_Media']),
            formatar_moeda_brasileira(forn['Valor_Total']),
            formatar_percentual_brasileiro(forn['Perc_Ate_44']),
            formatar_percentual_brasileiro(forn['Perc_45_70']),
            formatar_percentual_brasileiro(forn['Perc_Acima_71']),
            forn['Recomendacao']
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws_fornecedores.cell(row=row, column=col, value=valor)
            
            # Colorir baseado na recomenda√ß√£o
            if forn['Recomendacao'] == "‚ùå REJEITAR":
                cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
            elif forn['Recomendacao'] == "‚ö†Ô∏è REVISAR":
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
    
    # Ajustar larguras
    larguras = [40, 12, 12, 12, 15, 18, 12, 15, 12, 15]
    for col, largura in enumerate(larguras, 1):
        from openpyxl.utils import get_column_letter
        ws_fornecedores.column_dimensions[get_column_letter(col)].width = largura
    
    # 3. ABA DETALHAMENTO POR MERCADORIA
    ws_mercadorias = wb.create_sheet("üõçÔ∏è Detalhes por Mercadoria")
    
    # Cabe√ßalhos
    headers_merc = [
        'Carga', 'Pedido', 'Fornecedor', 'Filial', 'C√≥digo', 'Mercadoria', 
        'Quantidade Entrega', 'Saldo Pedido', 'Cobertura Atual', 
        'Nota Fiscal', 'Faixa Cobertura', 'Observa√ß√£o'
    ]
    
    for col, header in enumerate(headers_merc, 1):
        cell = ws_mercadorias.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ordenar dados por cobertura (mais cr√≠ticos primeiro)
    df_ordenado = df.sort_values('Cobertura Atual', ascending=False)
    
    # Preencher dados linha por linha com formata√ß√£o brasileira
    for row, (_, item) in enumerate(df_ordenado.iterrows(), 2):
        
        # Determinar faixa de cobertura
        cobertura = item['Cobertura Atual']
        if cobertura <= 44:
            faixa = "‚úÖ At√© 44 dias"
            cor = "E6F3E6"
            obs = "OK para aprova√ß√£o"
        elif cobertura <= 70:
            faixa = "‚ö†Ô∏è 45-70 dias"
            cor = "FFF2CC"
            obs = "Aten√ß√£o - revisar necessidade"
        else:
            faixa = "‚ùå Acima 71 dias"
            cor = "FFD6D6"
            obs = "CR√çTICO - considerar rejei√ß√£o"
        
        dados = [
            item['Carga'],
            item['Pedido'],
            item['Fornecedor'][:25],
            item['Filial'],
            item['C√≥d.'],
            item['Mercadoria'][:40],
            formatar_numero_brasileiro(item['Quantidade<br />Entrega']),
            formatar_moeda_brasileira(item['Saldo Pedido']),
            formatar_numero_brasileiro(cobertura),
            item['Nota Fiscal'] if pd.notna(item['Nota Fiscal']) else 'Sem NF',
            faixa,
            obs
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws_mercadorias.cell(row=row, column=col, value=valor)
            cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar larguras
    larguras_merc = [10, 12, 25, 20, 12, 40, 15, 15, 12, 15, 15, 30]
    for col, largura in enumerate(larguras_merc, 1):
        from openpyxl.utils import get_column_letter
        ws_mercadorias.column_dimensions[get_column_letter(col)].width = largura
    
    # 4. ABA FAIXAS POR FILIAL
    criar_aba_faixas_por_filial(wb, df)
    
    # 5. ABA FAIXAS POR FORNECEDOR E FILIAL
    criar_aba_faixas_fornecedor_filial(wb, df)

    # 6. ABA DISTRIBUI√á√ÉO POR VALOR
    criar_aba_distribuicao_valor(wb, df)
    
    # Salvar arquivo tempor√°rio
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

def gerar_resumo_analise(df):
    """
    Gera resumo da an√°lise para resposta JSON
    """
    
    total_fornecedores = df['Fornecedor'].nunique()
    total_itens = len(df)
    valor_total = df['Saldo Pedido'].sum()
    cobertura_media = df['Cobertura Atual'].mean()
    
    # An√°lise por faixas
    ate_44 = len(df[df['Cobertura Atual'] <= 44])
    entre_45_70 = len(df[(df['Cobertura Atual'] >= 45) & (df['Cobertura Atual'] <= 70)])
    acima_71 = len(df[df['Cobertura Atual'] >= 71])
    
    perc_ate_44 = (ate_44 / total_itens) * 100
    perc_45_70 = (entre_45_70 / total_itens) * 100
    perc_acima_71 = (acima_71 / total_itens) * 100
    
    # An√°lise de recomenda√ß√µes
    fornecedores_aprovar = 0
    fornecedores_revisar = 0
    fornecedores_rejeitar = 0
    valor_rejeitar = 0
    
    for fornecedor in df['Fornecedor'].unique():
        dados_forn = df[df['Fornecedor'] == fornecedor]
        cobertura_media_forn = dados_forn['Cobertura Atual'].mean()
        acima_71_forn = len(dados_forn[dados_forn['Cobertura Atual'] >= 71])
        perc_acima_71_forn = (acima_71_forn / len(dados_forn)) * 100
        valor_forn = dados_forn['Saldo Pedido'].sum()
        
        if perc_acima_71_forn > 50 or cobertura_media_forn > 100:
            fornecedores_rejeitar += 1
            valor_rejeitar += valor_forn
        elif perc_acima_71_forn > 25 or cobertura_media_forn > 70:
            fornecedores_revisar += 1
        else:
            fornecedores_aprovar += 1
    
    # Distribui√ß√£o por filial
    filiais_info = []
    for filial in df['Filial'].unique():
        dados_filial = df[df['Filial'] == filial]
        filiais_info.append({
            'nome': filial,
            'itens': len(dados_filial),
            'cobertura_media': dados_filial['Cobertura Atual'].mean(),
            'valor': dados_filial['Saldo Pedido'].sum()
        })
    
    resumo = {
        'metricas_gerais': {
            'total_fornecedores': total_fornecedores,
            'total_itens': total_itens,
            'valor_total': valor_total,
            'cobertura_media': cobertura_media
        },
        'distribuicao_faixas': {
            'ate_44_dias': {'quantidade': ate_44, 'percentual': perc_ate_44},
            'entre_45_70_dias': {'quantidade': entre_45_70, 'percentual': perc_45_70},
            'acima_71_dias': {'quantidade': acima_71, 'percentual': perc_acima_71}
        },
        'recomendacoes': {
            'aprovar': fornecedores_aprovar,
            'revisar': fornecedores_revisar,
            'rejeitar': fornecedores_rejeitar,
            'economia_potencial': valor_rejeitar
        },
        'filiais': filiais_info
    }
    
    return resumo

def criar_aba_faixas_por_filial(wb, df):
    """
    Cria aba com an√°lise detalhada de faixas por filial
    """
    
    ws = wb.create_sheet("üìç Faixas por Filial")
    
    # T√≠tulo principal
    ws.cell(row=1, column=1, value="AN√ÅLISE DE FAIXAS DE COBERTURA POR FILIAL").font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:J1')
    
    # Cabe√ßalhos
    headers = [
        'Filial', 'Total Itens', 'Valor Total (R$)', 'Cobertura M√©dia',
        'At√© 44 dias', '% At√© 44', 'Entre 45-70 dias', '% 45-70',
        'Acima 71 dias', '% Acima 71'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Analisar cada filial
    filiais_analise = []
    
    for filial in df['Filial'].unique():
        dados_filial = df[df['Filial'] == filial]
        total_itens = len(dados_filial)
        valor_total = dados_filial['Saldo Pedido'].sum()
        cobertura_media = dados_filial['Cobertura Atual'].mean()
        
        # Distribui√ß√£o por faixas
        ate_44 = len(dados_filial[dados_filial['Cobertura Atual'] <= 44])
        entre_45_70 = len(dados_filial[(dados_filial['Cobertura Atual'] >= 45) & (dados_filial['Cobertura Atual'] <= 70)])
        acima_71 = len(dados_filial[dados_filial['Cobertura Atual'] >= 71])
        
        perc_ate_44 = (ate_44 / total_itens) * 100
        perc_45_70 = (entre_45_70 / total_itens) * 100
        perc_acima_71 = (acima_71 / total_itens) * 100
        
        filiais_analise.append({
            'filial': filial,
            'total_itens': total_itens,
            'valor_total': valor_total,
            'cobertura_media': cobertura_media,
            'ate_44': ate_44,
            'perc_ate_44': perc_ate_44,
            'entre_45_70': entre_45_70,
            'perc_45_70': perc_45_70,
            'acima_71': acima_71,
            'perc_acima_71': perc_acima_71
        })
    
    # Ordenar por % acima de 71 dias (mais cr√≠ticos primeiro)
    filiais_analise.sort(key=lambda x: x['perc_acima_71'], reverse=True)
    
    # Preencher dados com formata√ß√£o brasileira
    for row, filial_data in enumerate(filiais_analise, 4):
        dados = [
            filial_data['filial'],
            formatar_numero_brasileiro(filial_data['total_itens']),
            formatar_moeda_brasileira(filial_data['valor_total']),
            formatar_numero_brasileiro(filial_data['cobertura_media']),
            formatar_numero_brasileiro(filial_data['ate_44']),
            formatar_percentual_brasileiro(filial_data['perc_ate_44']),
            formatar_numero_brasileiro(filial_data['entre_45_70']),
            formatar_percentual_brasileiro(filial_data['perc_45_70']),
            formatar_numero_brasileiro(filial_data['acima_71']),
            formatar_percentual_brasileiro(filial_data['perc_acima_71'])
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            
            # Colorir baseado na criticidade
            if filial_data['perc_acima_71'] > 50:
                cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
            elif filial_data['perc_acima_71'] > 25:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adicionar detalhamento por fornecedor dentro de cada filial
    row_atual = len(filiais_analise) + 6
    
    for filial_data in filiais_analise:
        filial = filial_data['filial']
        dados_filial = df[df['Filial'] == filial]
        
        # T√≠tulo da filial
        ws.cell(row=row_atual, column=1, value=f"DETALHAMENTO - {filial}").font = Font(bold=True, size=12)
        ws.cell(row=row_atual, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.merge_cells(f'A{row_atual}:J{row_atual}')
        
        row_atual += 2
        
        # Cabe√ßalhos do detalhamento
        headers_det = ['Fornecedor', 'Itens', 'At√© 44d', '% 44d', '45-70d', '% 45-70d', 'Acima 71d', '% 71d', 'Cobertura M√©dia']
        for col, header in enumerate(headers_det, 1):
            cell = ws.cell(row=row_atual, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        row_atual += 1
        
        # Analisar fornecedores da filial
        for fornecedor in dados_filial['Fornecedor'].unique():
            dados_forn_filial = dados_filial[dados_filial['Fornecedor'] == fornecedor]
            total_itens_forn = len(dados_forn_filial)
            cobertura_media_forn = dados_forn_filial['Cobertura Atual'].mean()
            
            ate_44_forn = len(dados_forn_filial[dados_forn_filial['Cobertura Atual'] <= 44])
            entre_45_70_forn = len(dados_forn_filial[(dados_forn_filial['Cobertura Atual'] >= 45) & (dados_forn_filial['Cobertura Atual'] <= 70)])
            acima_71_forn = len(dados_forn_filial[dados_forn_filial['Cobertura Atual'] >= 71])
            
            perc_ate_44_forn = (ate_44_forn / total_itens_forn) * 100
            perc_45_70_forn = (entre_45_70_forn / total_itens_forn) * 100
            perc_acima_71_forn = (acima_71_forn / total_itens_forn) * 100
            
            dados_forn = [
                fornecedor[:25],
                formatar_numero_brasileiro(total_itens_forn),
                formatar_numero_brasileiro(ate_44_forn),
                formatar_percentual_brasileiro(perc_ate_44_forn),
                formatar_numero_brasileiro(entre_45_70_forn),
                formatar_percentual_brasileiro(perc_45_70_forn),
                formatar_numero_brasileiro(acima_71_forn),
                formatar_percentual_brasileiro(perc_acima_71_forn),
                formatar_numero_brasileiro(cobertura_media_forn)
            ]
            
            for col, valor in enumerate(dados_forn, 1):
                cell = ws.cell(row=row_atual, column=col, value=valor)
                
                # Colorir baseado na criticidade do fornecedor
                if perc_acima_71_forn > 50:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif perc_acima_71_forn > 25:
                    cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            
            row_atual += 1
        
        row_atual += 2
    
    # Ajustar larguras
    larguras = [25, 10, 15, 10, 10, 10, 10, 10, 15]
    for col, largura in enumerate(larguras, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = largura

def criar_aba_faixas_fornecedor_filial(wb, df):

    # 6. ABA DISTRIBUI√á√ÉO POR VALOR
    """
    Cria aba com an√°lise detalhada de faixas por fornecedor e filial (TODOS)
    """
    
    ws = wb.create_sheet("üè≠ Faixas Fornecedor x Filial")
    
    # T√≠tulo principal
    ws.cell(row=1, column=1, value="AN√ÅLISE DE FAIXAS POR FORNECEDOR E FILIAL - TODOS").font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    ws.merge_cells('A1:K1')
    
    # Cabe√ßalhos
    headers = [
        'Fornecedor', 'Filial', 'Total Itens', 'Valor Total (R$)', 'Cobertura M√©dia',
        'At√© 44 dias', '% At√© 44', 'Entre 45-70 dias', '% 45-70',
        'Acima 71 dias', '% Acima 71'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Analisar cada combina√ß√£o fornecedor-filial
    combinacoes_analise = []
    
    for fornecedor in df['Fornecedor'].unique():
        dados_forn = df[df['Fornecedor'] == fornecedor]
        
        for filial in dados_forn['Filial'].unique():
            dados_forn_filial = dados_forn[dados_forn['Filial'] == filial]
            
            if len(dados_forn_filial) == 0:
                continue
            
            total_itens = len(dados_forn_filial)
            valor_total = dados_forn_filial['Saldo Pedido'].sum()
            cobertura_media = dados_forn_filial['Cobertura Atual'].mean()
            
            # Distribui√ß√£o por faixas
            ate_44 = len(dados_forn_filial[dados_forn_filial['Cobertura Atual'] <= 44])
            entre_45_70 = len(dados_forn_filial[(dados_forn_filial['Cobertura Atual'] >= 45) & (dados_forn_filial['Cobertura Atual'] <= 70)])
            acima_71 = len(dados_forn_filial[dados_forn_filial['Cobertura Atual'] >= 71])
            
            perc_ate_44 = (ate_44 / total_itens) * 100
            perc_45_70 = (entre_45_70 / total_itens) * 100
            perc_acima_71 = (acima_71 / total_itens) * 100
            
            combinacoes_analise.append({
                'fornecedor': fornecedor,
                'filial': filial,
                'total_itens': total_itens,
                'valor_total': valor_total,
                'cobertura_media': cobertura_media,
                'ate_44': ate_44,
                'perc_ate_44': perc_ate_44,
                'entre_45_70': entre_45_70,
                'perc_45_70': perc_45_70,
                'acima_71': acima_71,
                'perc_acima_71': perc_acima_71
            })
    
    # Ordenar por % acima de 71 dias (mais cr√≠ticos primeiro)
    combinacoes_analise.sort(key=lambda x: x['perc_acima_71'], reverse=True)
    
    # Preencher dados com formata√ß√£o brasileira
    for row, comb_data in enumerate(combinacoes_analise, 4):
        dados = [
            comb_data['fornecedor'][:25],
            comb_data['filial'],
            formatar_numero_brasileiro(comb_data['total_itens']),
            formatar_moeda_brasileira(comb_data['valor_total']),
            formatar_numero_brasileiro(comb_data['cobertura_media']),
            formatar_numero_brasileiro(comb_data['ate_44']),
            formatar_percentual_brasileiro(comb_data['perc_ate_44']),
            formatar_numero_brasileiro(comb_data['entre_45_70']),
            formatar_percentual_brasileiro(comb_data['perc_45_70']),
            formatar_numero_brasileiro(comb_data['acima_71']),
            formatar_percentual_brasileiro(comb_data['perc_acima_71'])
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            
            # Colorir baseado na criticidade
            if comb_data['perc_acima_71'] > 50:
                cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
            elif comb_data['perc_acima_71'] > 25:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar larguras
    larguras = [25, 20, 10, 15, 12, 10, 10, 12, 10, 10, 15]
    for col, largura in enumerate(larguras, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = largura

@analise_bp.route('/upload', methods=['POST'])
def upload_arquivo():
    """
    Endpoint para upload e processamento do arquivo
    """
    
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        try:
            file.save(filepath)
            
            # Processar arquivo
            output_file, resultado = processar_arquivo_cargas(filepath)
            
            if output_file is None:
                return jsonify({'error': resultado}), 400
            
            # Salvar caminho do arquivo gerado na sess√£o ou cache
            # Para simplificar, vamos usar um nome baseado no timestamp
            output_filename = f"analise_{timestamp}.xlsx"
            final_output_path = os.path.join(UPLOAD_FOLDER, output_filename)
            
            # Mover arquivo tempor√°rio para local permanente
            import shutil
            shutil.move(output_file, final_output_path)
            
            # Limpar arquivo original
            os.remove(filepath)
            
            return jsonify({
                'success': True,
                'message': 'Arquivo processado com sucesso!',
                'download_url': f'/api/analise/download/{output_filename}',
                'resumo': resultado
            })
            
        except Exception as e:
            return jsonify({'error': f'Erro ao processar arquivo: {str(e)}'}), 500
    
    return jsonify({'error': 'Tipo de arquivo n√£o permitido'}), 400

@analise_bp.route('/download/<filename>')
def download_arquivo(filename):
    """
    Endpoint para download do arquivo de an√°lise
    """
    try:
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Arquivo n√£o encontrado'}), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=f"Analise_Cargas_Aprovacao_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Erro ao fazer download: {str(e)}'}), 500
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment

def criar_aba_distribuicao_valor(wb, df):
    """
    Cria aba com an√°lise de distribui√ß√£o por faixas de valor
    """
    
    ws = wb.create_sheet("üí∞ Distribui√ß√£o por Valor")
    
    # T√≠tulo principal
    ws.cell(row=1, column=1, value="AN√ÅLISE DE DISTRIBUI√á√ÉO POR FAIXAS DE VALOR").font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    ws.merge_cells('A1:H1')
    
    # Definir faixas de valor
    faixas_valor = [
        {"nome": "At√© R$ 100", "min": 0, "max": 100},
        {"nome": "R$ 101 - R$ 500", "min": 101, "max": 500},
        {"nome": "R$ 501 - R$ 1.000", "min": 501, "max": 1000},
        {"nome": "R$ 1.001 - R$ 2.500", "min": 1001, "max": 2500},
        {"nome": "R$ 2.501 - R$ 5.000", "min": 2501, "max": 5000},
        {"nome": "R$ 5.001 - R$ 10.000", "min": 5001, "max": 10000},
        {"nome": "Acima de R$ 10.000", "min": 10001, "max": float('inf')}
    ]
    
    # Cabe√ßalhos
    headers = [
        'Faixa de Valor', 'Quantidade Itens', '% do Total', 'Valor Total (R$)', 
        '% do Valor Total', 'Cobertura M√©dia', 'Recomenda√ß√£o', 'Observa√ß√£o'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Calcular totais gerais
    total_itens = len(df)
    valor_total_geral = df['Saldo Pedido'].sum()
    
    # Analisar cada faixa
    faixas_analise = []
    
    for faixa in faixas_valor:
        if faixa["max"] == float('inf'):
            dados_faixa = df[df['Saldo Pedido'] >= faixa["min"]]
        else:
            dados_faixa = df[(df['Saldo Pedido'] >= faixa["min"]) & (df['Saldo Pedido'] <= faixa["max"])]
        
        quantidade = len(dados_faixa)
        if quantidade == 0:
            continue
            
        valor_faixa = dados_faixa['Saldo Pedido'].sum()
        cobertura_media = dados_faixa['Cobertura Atual'].mean()
        
        perc_quantidade = (quantidade / total_itens) * 100
        perc_valor = (valor_faixa / valor_total_geral) * 100
        
        # Determinar recomenda√ß√£o baseada na cobertura m√©dia
        if cobertura_media <= 44:
            recomendacao = "‚úÖ APROVAR"
            observacao = "Faixa com boa rotatividade"
        elif cobertura_media <= 70:
            recomendacao = "‚ö†Ô∏è REVISAR"
            observacao = "Aten√ß√£o √† cobertura m√©dia"
        else:
            recomendacao = "‚ùå REJEITAR"
            observacao = "Cobertura alta - risco de estoque parado"
        
        faixas_analise.append({
            'nome': faixa["nome"],
            'quantidade': quantidade,
            'perc_quantidade': perc_quantidade,
            'valor_total': valor_faixa,
            'perc_valor': perc_valor,
            'cobertura_media': cobertura_media,
            'recomendacao': recomendacao,
            'observacao': observacao
        })
    
    # Preencher dados
    for row, faixa_data in enumerate(faixas_analise, 4):
        dados = [
            faixa_data['nome'],
            f"{faixa_data['quantidade']:,}".replace(",", "."),
            f"{faixa_data['perc_quantidade']:.1f}%".replace(".", ","),
            f"R$ {faixa_data['valor_total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            f"{faixa_data['perc_valor']:.1f}%".replace(".", ","),
            f"{faixa_data['cobertura_media']:.1f}".replace(".", ","),
            faixa_data['recomendacao'],
            faixa_data['observacao']
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            
            # Colorir baseado na recomenda√ß√£o
            if faixa_data['recomendacao'] == "‚ùå REJEITAR":
                cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
            elif faixa_data['recomendacao'] == "‚ö†Ô∏è REVISAR":
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adicionar linha de totais
    row_total = len(faixas_analise) + 6
    ws.cell(row=row_total, column=1, value="TOTAL GERAL").font = Font(bold=True)
    ws.cell(row=row_total, column=1).fill = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
    
    ws.cell(row=row_total, column=2, value=f"{total_itens:,}".replace(",", ".")).font = Font(bold=True)
    ws.cell(row=row_total, column=2).fill = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
    
    ws.cell(row=row_total, column=3, value="100,0%").font = Font(bold=True)
    ws.cell(row=row_total, column=3).fill = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
    
    ws.cell(row=row_total, column=4, value=f"R$ {valor_total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")).font = Font(bold=True)
    ws.cell(row=row_total, column=4).fill = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
    
    ws.cell(row=row_total, column=5, value="100,0%").font = Font(bold=True)
    ws.cell(row=row_total, column=5).fill = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
    
    # Adicionar an√°lise detalhada por fornecedor em cada faixa
    row_atual = row_total + 3
    
    for faixa_data in faixas_analise:
        faixa = next(f for f in faixas_valor if f["nome"] == faixa_data['nome'])
        
        if faixa["max"] == float('inf'):
            dados_faixa = df[df['Saldo Pedido'] >= faixa["min"]]
        else:
            dados_faixa = df[(df['Saldo Pedido'] >= faixa["min"]) & (df['Saldo Pedido'] <= faixa["max"])]
        
        if len(dados_faixa) == 0:
            continue
        
        # T√≠tulo da faixa
        ws.cell(row=row_atual, column=1, value=f"DETALHAMENTO - {faixa_data['nome']}").font = Font(bold=True, size=12)
        ws.cell(row=row_atual, column=1).fill = PatternFill(start_color="F8C471", end_color="F8C471", fill_type="solid")
        ws.merge_cells(f'A{row_atual}:H{row_atual}')
        
        row_atual += 2
        
        # Cabe√ßalhos do detalhamento
        headers_det = ['Fornecedor', 'Filial', 'Itens', 'Valor Total', 'Cobertura M√©dia', 'Maior Valor', 'Menor Valor', 'Recomenda√ß√£o']
        for col, header in enumerate(headers_det, 1):
            cell = ws.cell(row=row_atual, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        row_atual += 1
        
        # Analisar fornecedores na faixa
        fornecedores_faixa = []
        for fornecedor in dados_faixa['Fornecedor'].unique():
            dados_forn_faixa = dados_faixa[dados_faixa['Fornecedor'] == fornecedor]
            
            for filial in dados_forn_faixa['Filial'].unique():
                dados_forn_filial = dados_forn_faixa[dados_forn_faixa['Filial'] == filial]
                
                if len(dados_forn_filial) == 0:
                    continue
                
                valor_total_forn = dados_forn_filial['Saldo Pedido'].sum()
                cobertura_media_forn = dados_forn_filial['Cobertura Atual'].mean()
                maior_valor = dados_forn_filial['Saldo Pedido'].max()
                menor_valor = dados_forn_filial['Saldo Pedido'].min()
                
                # Recomenda√ß√£o espec√≠fica
                if cobertura_media_forn <= 44:
                    rec_forn = "‚úÖ APROVAR"
                elif cobertura_media_forn <= 70:
                    rec_forn = "‚ö†Ô∏è REVISAR"
                else:
                    rec_forn = "‚ùå REJEITAR"
                
                fornecedores_faixa.append({
                    'fornecedor': fornecedor,
                    'filial': filial,
                    'itens': len(dados_forn_filial),
                    'valor_total': valor_total_forn,
                    'cobertura_media': cobertura_media_forn,
                    'maior_valor': maior_valor,
                    'menor_valor': menor_valor,
                    'recomendacao': rec_forn
                })
        
        # Ordenar por valor total (maiores primeiro)
        fornecedores_faixa.sort(key=lambda x: x['valor_total'], reverse=True)
        
        for forn_data in fornecedores_faixa:
            dados_forn = [
                forn_data['fornecedor'][:20],
                forn_data['filial'][:15],
                f"{forn_data['itens']:,}".replace(",", "."),
                f"R$ {forn_data['valor_total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"{forn_data['cobertura_media']:.1f}".replace(".", ","),
                f"R$ {forn_data['maior_valor']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"R$ {forn_data['menor_valor']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                forn_data['recomendacao']
            ]
            
            for col, valor in enumerate(dados_forn, 1):
                cell = ws.cell(row=row_atual, column=col, value=valor)
                
                # Colorir baseado na recomenda√ß√£o
                if forn_data['recomendacao'] == "‚ùå REJEITAR":
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif forn_data['recomendacao'] == "‚ö†Ô∏è REVISAR":
                    cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            
            row_atual += 1
        
        row_atual += 2
    
    # Ajustar larguras
    larguras = [25, 15, 12, 18, 15, 15, 15, 20]
    for col, largura in enumerate(larguras, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = largura

# Teste da fun√ß√£o
if __name__ == "__main__":
    print("Fun√ß√£o de an√°lise de distribui√ß√£o por valor criada com sucesso!")
