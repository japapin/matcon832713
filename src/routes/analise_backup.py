import os
import pandas as pd
import numpy as np
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tempfile
import io

analise_bp = Blueprint('analise', __name__)

UPLOAD_FOLDER = '/tmp/uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Criar pasta de upload se não existir
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def processar_arquivo_cargas(filepath):
    """
    Processa arquivo de cargas e gera análise completa
    """
    try:
        # Carregar dados originais
        df_original = pd.read_excel(filepath, sheet_name='Agenda Recebimento')
        
        # Filtrar apenas "Em Aprovação"
        df_aprovacao = df_original[df_original['Status'] == 'Em Aprovação'].copy()
        
        if len(df_aprovacao) == 0:
            return None, "Nenhum registro encontrado com status 'Em Aprovação'"
        
        # Converter dados numéricos
        df_aprovacao['Cobertura Atual'] = pd.to_numeric(df_aprovacao['Cobertura Atual'], errors='coerce')
        df_aprovacao['Saldo Pedido'] = pd.to_numeric(df_aprovacao['Saldo Pedido'], errors='coerce')
        df_aprovacao['Quantidade<br />Entrega'] = pd.to_numeric(df_aprovacao['Quantidade<br />Entrega'], errors='coerce')
        
        # Limpar dados
        df_clean = df_aprovacao[
            df_aprovacao['Cobertura Atual'].notna() & 
            df_aprovacao['Fornecedor'].notna() &
            df_aprovacao['Filial'].notna() &
            df_aprovacao['Mercadoria'].notna()
        ].copy()
        
        if len(df_clean) == 0:
            return None, "Nenhum registro válido encontrado após limpeza dos dados"
        
        # Gerar arquivo Excel
        output_file = gerar_excel_analise(df_clean)
        
        # Gerar resumo para resposta
        resumo = gerar_resumo_analise(df_clean)
        
        return output_file, resumo
        
    except Exception as e:
        return None, f"Erro ao processar arquivo: {str(e)}"

def gerar_excel_analise(df):
    """
    Gera arquivo Excel com análise completa
    """
    
    # Criar workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # 1. ABA RESUMO EXECUTIVO
    ws_resumo = wb.create_sheet("📊 Resumo Executivo")
    
    # Calcular métricas principais
    total_fornecedores = df['Fornecedor'].nunique()
    total_cargas = df['Carga'].nunique()
    total_itens = len(df)
    total_filiais = df['Filial'].nunique()
    valor_total = df['Saldo Pedido'].sum()
    cobertura_media_geral = df['Cobertura Atual'].mean()
    
    # Análise por faixas
    ate_44 = len(df[df['Cobertura Atual'] <= 44])
    entre_45_70 = len(df[(df['Cobertura Atual'] >= 45) & (df['Cobertura Atual'] <= 70)])
    acima_71 = len(df[df['Cobertura Atual'] >= 71])
    
    perc_ate_44 = (ate_44 / total_itens) * 100
    perc_45_70 = (entre_45_70 / total_itens) * 100
    perc_acima_71 = (acima_71 / total_itens) * 100
    
    # Dados do resumo
    dados_resumo = [
        ["ANÁLISE DE CARGAS EM APROVAÇÃO", ""],
        [f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ""],
        ["", ""],
        ["MÉTRICAS GERAIS", ""],
        ["Total de Fornecedores", total_fornecedores],
        ["Total de Cargas", total_cargas],
        ["Total de Itens", f"{total_itens:,}"],
        ["Total de Filiais", total_filiais],
        ["Valor Total", f"R$ {valor_total:,.2f}"],
        ["Cobertura Média Geral", f"{cobertura_media_geral:.1f} dias"],
        ["", ""],
        ["DISTRIBUIÇÃO POR FAIXAS DE COBERTURA", ""],
        ["Até 44 dias", f"{ate_44:,} itens ({perc_ate_44:.1f}%)"],
        ["Entre 45-70 dias", f"{entre_45_70:,} itens ({perc_45_70:.1f}%)"],
        ["Acima de 71 dias", f"{acima_71:,} itens ({perc_acima_71:.1f}%)"],
        ["", ""],
        ["DISTRIBUIÇÃO POR FILIAL", ""],
    ]
    
    # Adicionar dados por filial
    for filial in df['Filial'].unique():
        dados_filial = df[df['Filial'] == filial]
        cobertura_filial = dados_filial['Cobertura Atual'].mean()
        valor_filial = dados_filial['Saldo Pedido'].sum()
        dados_resumo.append([filial, f"{len(dados_filial):,} itens, {cobertura_filial:.1f} dias, R$ {valor_filial:,.0f}"])
    
    # Preencher dados
    for row, (label, valor) in enumerate(dados_resumo, 1):
        cell_a = ws_resumo.cell(row=row, column=1, value=label)
        cell_b = ws_resumo.cell(row=row, column=2, value=valor)
        
        # Formatação
        if "ANÁLISE DE CARGAS" in str(label):
            cell_a.font = Font(bold=True, size=16, color="FFFFFF")
            cell_a.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        elif label in ["MÉTRICAS GERAIS", "DISTRIBUIÇÃO POR FAIXAS DE COBERTURA", "DISTRIBUIÇÃO POR FILIAL"]:
            cell_a.font = Font(bold=True, size=12)
            cell_a.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_b.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar larguras
    ws_resumo.column_dimensions['A'].width = 40
    ws_resumo.column_dimensions['B'].width = 30
    
    # 2. ABA ANÁLISE POR FORNECEDOR
    ws_fornecedores = wb.create_sheet("🏭 Análise por Fornecedor")
    
    # Cabeçalhos
    headers = [
        'Fornecedor', 'Total Itens', 'Total Cargas', 'Filiais Atendidas', 
        'Cobertura Média', 'Valor Total (R$)', '% Até 44 dias', 
        '% Entre 45-70 dias', '% Acima 71 dias', 'Recomendação'
    ]
    
    # Criar cabeçalho
    for col, header in enumerate(headers, 1):
        cell = ws_fornecedores.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Analisar cada fornecedor
    fornecedores_analise = []
    
    for fornecedor in df['Fornecedor'].unique():
        dados_forn = df[df['Fornecedor'] == fornecedor]
        
        total_itens = len(dados_forn)
        total_cargas = dados_forn['Carga'].nunique()
        filiais = dados_forn['Filial'].nunique()
        cobertura_media = dados_forn['Cobertura Atual'].mean()
        valor_total = dados_forn['Saldo Pedido'].sum()
        
        # Distribuição por faixas
        ate_44 = len(dados_forn[dados_forn['Cobertura Atual'] <= 44])
        entre_45_70 = len(dados_forn[(dados_forn['Cobertura Atual'] >= 45) & (dados_forn['Cobertura Atual'] <= 70)])
        acima_71 = len(dados_forn[dados_forn['Cobertura Atual'] >= 71])
        
        perc_ate_44 = (ate_44 / total_itens) * 100
        perc_45_70 = (entre_45_70 / total_itens) * 100
        perc_acima_71 = (acima_71 / total_itens) * 100
        
        # Recomendação
        if perc_acima_71 > 50 or cobertura_media > 100:
            recomendacao = "❌ REJEITAR"
        elif perc_acima_71 > 25 or cobertura_media > 70:
            recomendacao = "⚠️ REVISAR"
        else:
            recomendacao = "✅ APROVAR"
        
        fornecedores_analise.append({
            'Fornecedor': fornecedor,
            'Total_Itens': total_itens,
            'Total_Cargas': total_cargas,
            'Filiais': filiais,
            'Cobertura_Media': cobertura_media,
            'Valor_Total': valor_total,
            'Perc_Ate_44': perc_ate_44,
            'Perc_45_70': perc_45_70,
            'Perc_Acima_71': perc_acima_71,
            'Recomendacao': recomendacao
        })
    
    # Ordenar por cobertura média (mais críticos primeiro)
    fornecedores_analise.sort(key=lambda x: x['Cobertura_Media'], reverse=True)
    
    # Preencher dados
    for row, forn in enumerate(fornecedores_analise, 2):
        dados = [
            forn['Fornecedor'][:40],
            forn['Total_Itens'],
            forn['Total_Cargas'],
            forn['Filiais'],
            f"{forn['Cobertura_Media']:.1f}",
            f"R$ {forn['Valor_Total']:,.2f}",
            f"{forn['Perc_Ate_44']:.1f}%",
            f"{forn['Perc_45_70']:.1f}%",
            f"{forn['Perc_Acima_71']:.1f}%",
            forn['Recomendacao']
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws_fornecedores.cell(row=row, column=col, value=valor)
            
            # Colorir baseado na recomendação
            if forn['Recomendacao'] == "❌ REJEITAR":
                cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
            elif forn['Recomendacao'] == "⚠️ REVISAR":
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
    
    # Ajustar larguras
    larguras = [40, 12, 12, 12, 15, 18, 12, 15, 12, 15]
    for col, largura in enumerate(larguras, 1):
        from openpyxl.utils import get_column_letter
        ws_fornecedores.column_dimensions[get_column_letter(col)].width = largura
    
    # 3. ABA DETALHAMENTO POR MERCADORIA
    ws_mercadorias = wb.create_sheet("🛍️ Detalhes por Mercadoria")
    
    # Cabeçalhos
    headers_merc = [
        'Carga', 'Fornecedor', 'Filial', 'Código', 'Mercadoria', 
        'Quantidade Entrega', 'Saldo Pedido', 'Cobertura Atual', 
        'Faixa Cobertura', 'Observação'
    ]
    
    for col, header in enumerate(headers_merc, 1):
        cell = ws_mercadorias.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ordenar dados por cobertura (mais críticos primeiro)
    df_ordenado = df.sort_values('Cobertura Atual', ascending=False)
    
    # Preencher dados linha por linha
    for row, (_, item) in enumerate(df_ordenado.iterrows(), 2):
        
        # Determinar faixa de cobertura
        cobertura = item['Cobertura Atual']
        if cobertura <= 44:
            faixa = "✅ Até 44 dias"
            cor = "E6F3E6"
            obs = "OK para aprovação"
        elif cobertura <= 70:
            faixa = "⚠️ 45-70 dias"
            cor = "FFF2CC"
            obs = "Atenção - revisar necessidade"
        else:
            faixa = "❌ Acima 71 dias"
            cor = "FFD6D6"
            obs = "CRÍTICO - considerar rejeição"
        
        dados = [
            item['Carga'],
            item['Fornecedor'][:25],
            item['Filial'],
            item['Cód.'],
            item['Mercadoria'][:40],
            item['Quantidade<br />Entrega'],
            f"R$ {item['Saldo Pedido']:,.2f}",
            f"{cobertura:.1f}",
            faixa,
            obs
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws_mercadorias.cell(row=row, column=col, value=valor)
            cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar larguras
    larguras_merc = [10, 25, 20, 12, 40, 15, 15, 12, 15, 30]
    for col, largura in enumerate(larguras_merc, 1):
        from openpyxl.utils import get_column_letter
        ws_mercadorias.column_dimensions[get_column_letter(col)].width = largura
    
    # 4. ABA FAIXAS POR FILIAL
    criar_aba_faixas_por_filial(wb, df)
    
    # 5. ABA FAIXAS POR FORNECEDOR
    criar_aba_faixas_por_fornecedor(wb, df)
    
    # Salvar arquivo temporário
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

def gerar_resumo_analise(df):
    """
    Gera resumo da análise para resposta JSON
    """
    
    total_fornecedores = df['Fornecedor'].nunique()
    total_itens = len(df)
    valor_total = df['Saldo Pedido'].sum()
    cobertura_media = df['Cobertura Atual'].mean()
    
    # Análise por faixas
    ate_44 = len(df[df['Cobertura Atual'] <= 44])
    entre_45_70 = len(df[(df['Cobertura Atual'] >= 45) & (df['Cobertura Atual'] <= 70)])
    acima_71 = len(df[df['Cobertura Atual'] >= 71])
    
    perc_ate_44 = (ate_44 / total_itens) * 100
    perc_45_70 = (entre_45_70 / total_itens) * 100
    perc_acima_71 = (acima_71 / total_itens) * 100
    
    # Análise de recomendações
    fornecedores_aprovar = 0
    fornecedores_revisar = 0
    fornecedores_rejeitar = 0
    valor_rejeitar = 0
    
    for fornecedor in df['Fornecedor'].unique():
        dados_forn = df[df['Fornecedor'] == fornecedor]
        cobertura_media_forn = dados_forn['Cobertura Atual'].mean()
        acima_71_forn = len(dados_forn[dados_forn['Cobertura Atual'] >= 71])
        perc_acima_71_forn = (acima_71_forn / len(dados_forn)) * 100
        valor_forn = dados_forn['Saldo Pedido'].sum()
        
        if perc_acima_71_forn > 50 or cobertura_media_forn > 100:
            fornecedores_rejeitar += 1
            valor_rejeitar += valor_forn
        elif perc_acima_71_forn > 25 or cobertura_media_forn > 70:
            fornecedores_revisar += 1
        else:
            fornecedores_aprovar += 1
    
    # Distribuição por filial
    filiais_info = []
    for filial in df['Filial'].unique():
        dados_filial = df[df['Filial'] == filial]
        filiais_info.append({
            'nome': filial,
            'itens': len(dados_filial),
            'cobertura_media': dados_filial['Cobertura Atual'].mean(),
            'valor': dados_filial['Saldo Pedido'].sum()
        })
    
    resumo = {
        'metricas_gerais': {
            'total_fornecedores': total_fornecedores,
            'total_itens': total_itens,
            'valor_total': valor_total,
            'cobertura_media': cobertura_media
        },
        'distribuicao_faixas': {
            'ate_44_dias': {'quantidade': ate_44, 'percentual': perc_ate_44},
            'entre_45_70_dias': {'quantidade': entre_45_70, 'percentual': perc_45_70},
            'acima_71_dias': {'quantidade': acima_71, 'percentual': perc_acima_71}
        },
        'recomendacoes': {
            'aprovar': fornecedores_aprovar,
            'revisar': fornecedores_revisar,
            'rejeitar': fornecedores_rejeitar,
            'economia_potencial': valor_rejeitar
        },
        'filiais': filiais_info
    }
    
    return resumo

@analise_bp.route('/upload', methods=['POST'])
def upload_arquivo():
    """
    Endpoint para upload e processamento do arquivo
    """
    
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        try:
            file.save(filepath)
            
            # Processar arquivo
            output_file, resultado = processar_arquivo_cargas(filepath)
            
            if output_file is None:
                return jsonify({'error': resultado}), 400
            
            # Salvar caminho do arquivo gerado na sessão ou cache
            # Para simplificar, vamos usar um nome baseado no timestamp
            output_filename = f"analise_{timestamp}.xlsx"
            final_output_path = os.path.join(UPLOAD_FOLDER, output_filename)
            
            # Mover arquivo temporário para local permanente
            import shutil
            shutil.move(output_file, final_output_path)
            
            # Limpar arquivo original
            os.remove(filepath)
            
            return jsonify({
                'success': True,
                'message': 'Arquivo processado com sucesso!',
                'download_url': f'/api/analise/download/{output_filename}',
                'resumo': resultado
            })
            
        except Exception as e:
            return jsonify({'error': f'Erro ao processar arquivo: {str(e)}'}), 500
    
    return jsonify({'error': 'Tipo de arquivo não permitido. Use apenas .xlsx ou .xls'}), 400

@analise_bp.route('/download/<filename>')
def download_arquivo(filename):
    """
    Endpoint para download do arquivo processado
    """
    
    try:
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Arquivo não encontrado'}), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=f"Analise_Cargas_Aprovacao_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Erro ao fazer download: {str(e)}'}), 500

@analise_bp.route('/status')
def status():
    """
    Endpoint para verificar status da API
    """
    return jsonify({
        'status': 'online',
        'message': 'API de Análise de Cargas funcionando',
        'timestamp': datetime.now().isoformat()
    })


def criar_aba_faixas_por_filial(wb, df):
    """
    Cria aba com análise detalhada de faixas por filial
    """
    
    ws = wb.create_sheet("📍 Faixas por Filial")
    
    # Título principal
    ws.cell(row=1, column=1, value="ANÁLISE DE FAIXAS DE COBERTURA POR FILIAL").font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:J1')
    
    # Cabeçalhos
    headers = [
        'Filial', 'Total Itens', 'Valor Total (R$)', 'Cobertura Média',
        'Até 44 dias', '% Até 44', 'Entre 45-70 dias', '% 45-70',
        'Acima 71 dias', '% Acima 71'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Analisar cada filial
    filiais_analise = []
    
    for filial in df['Filial'].unique():
        dados_filial = df[df['Filial'] == filial]
        total_itens = len(dados_filial)
        valor_total = dados_filial['Saldo Pedido'].sum()
        cobertura_media = dados_filial['Cobertura Atual'].mean()
        
        # Distribuição por faixas
        ate_44 = len(dados_filial[dados_filial['Cobertura Atual'] <= 44])
        entre_45_70 = len(dados_filial[(dados_filial['Cobertura Atual'] >= 45) & (dados_filial['Cobertura Atual'] <= 70)])
        acima_71 = len(dados_filial[dados_filial['Cobertura Atual'] >= 71])
        
        perc_ate_44 = (ate_44 / total_itens) * 100
        perc_45_70 = (entre_45_70 / total_itens) * 100
        perc_acima_71 = (acima_71 / total_itens) * 100
        
        filiais_analise.append({
            'filial': filial,
            'total_itens': total_itens,
            'valor_total': valor_total,
            'cobertura_media': cobertura_media,
            'ate_44': ate_44,
            'perc_ate_44': perc_ate_44,
            'entre_45_70': entre_45_70,
            'perc_45_70': perc_45_70,
            'acima_71': acima_71,
            'perc_acima_71': perc_acima_71
        })
    
    # Ordenar por % acima de 71 dias (mais críticos primeiro)
    filiais_analise.sort(key=lambda x: x['perc_acima_71'], reverse=True)
    
    # Preencher dados
    for row, filial_data in enumerate(filiais_analise, 4):
        dados = [
            filial_data['filial'],
            filial_data['total_itens'],
            f"R$ {filial_data['valor_total']:,.2f}",
            f"{filial_data['cobertura_media']:.1f}",
            filial_data['ate_44'],
            f"{filial_data['perc_ate_44']:.1f}%",
            filial_data['entre_45_70'],
            f"{filial_data['perc_45_70']:.1f}%",
            filial_data['acima_71'],
            f"{filial_data['perc_acima_71']:.1f}%"
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            
            # Colorir baseado na criticidade
            if filial_data['perc_acima_71'] > 50:
                cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
            elif filial_data['perc_acima_71'] > 25:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adicionar detalhamento por fornecedor dentro de cada filial
    row_atual = len(filiais_analise) + 6
    
    for filial_data in filiais_analise:
        filial = filial_data['filial']
        dados_filial = df[df['Filial'] == filial]
        
        # Título da filial
        ws.cell(row=row_atual, column=1, value=f"DETALHAMENTO - {filial}").font = Font(bold=True, size=12)
        ws.cell(row=row_atual, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.merge_cells(f'A{row_atual}:J{row_atual}')
        
        row_atual += 2
        
        # Cabeçalhos do detalhamento
        headers_det = ['Fornecedor', 'Itens', 'Até 44d', '% 44d', '45-70d', '% 45-70d', 'Acima 71d', '% 71d', 'Cobertura Média']
        for col, header in enumerate(headers_det, 1):
            cell = ws.cell(row=row_atual, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        row_atual += 1
        
        # Analisar fornecedores da filial
        for fornecedor in dados_filial['Fornecedor'].unique():
            dados_forn_filial = dados_filial[dados_filial['Fornecedor'] == fornecedor]
            total_itens_forn = len(dados_forn_filial)
            cobertura_media_forn = dados_forn_filial['Cobertura Atual'].mean()
            
            ate_44_forn = len(dados_forn_filial[dados_forn_filial['Cobertura Atual'] <= 44])
            entre_45_70_forn = len(dados_forn_filial[(dados_forn_filial['Cobertura Atual'] >= 45) & (dados_forn_filial['Cobertura Atual'] <= 70)])
            acima_71_forn = len(dados_forn_filial[dados_forn_filial['Cobertura Atual'] >= 71])
            
            perc_ate_44_forn = (ate_44_forn / total_itens_forn) * 100
            perc_45_70_forn = (entre_45_70_forn / total_itens_forn) * 100
            perc_acima_71_forn = (acima_71_forn / total_itens_forn) * 100
            
            dados_forn = [
                fornecedor[:25],
                total_itens_forn,
                ate_44_forn,
                f"{perc_ate_44_forn:.1f}%",
                entre_45_70_forn,
                f"{perc_45_70_forn:.1f}%",
                acima_71_forn,
                f"{perc_acima_71_forn:.1f}%",
                f"{cobertura_media_forn:.1f}"
            ]
            
            for col, valor in enumerate(dados_forn, 1):
                cell = ws.cell(row=row_atual, column=col, value=valor)
                
                # Colorir baseado na criticidade do fornecedor
                if perc_acima_71_forn > 50:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif perc_acima_71_forn > 25:
                    cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            
            row_atual += 1
        
        row_atual += 2
    
    # Ajustar larguras
    larguras = [25, 10, 15, 10, 10, 10, 10, 10, 15]
    for col, largura in enumerate(larguras, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = largura

def criar_aba_faixas_por_fornecedor(wb, df):
    """
    Cria aba com análise detalhada de faixas por fornecedor
    """
    
    ws = wb.create_sheet("🏭 Faixas por Fornecedor")
    
    # Título principal
    ws.cell(row=1, column=1, value="ANÁLISE DE FAIXAS DE COBERTURA POR FORNECEDOR").font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    ws.merge_cells('A1:K1')
    
    # Cabeçalhos
    headers = [
        'Fornecedor', 'Total Itens', 'Filiais', 'Valor Total (R$)', 'Cobertura Média',
        'Até 44 dias', '% Até 44', 'Entre 45-70 dias', '% 45-70',
        'Acima 71 dias', '% Acima 71'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Analisar cada fornecedor
    fornecedores_analise = []
    
    for fornecedor in df['Fornecedor'].unique():
        dados_forn = df[df['Fornecedor'] == fornecedor]
        total_itens = len(dados_forn)
        total_filiais = dados_forn['Filial'].nunique()
        valor_total = dados_forn['Saldo Pedido'].sum()
        cobertura_media = dados_forn['Cobertura Atual'].mean()
        
        # Distribuição por faixas
        ate_44 = len(dados_forn[dados_forn['Cobertura Atual'] <= 44])
        entre_45_70 = len(dados_forn[(dados_forn['Cobertura Atual'] >= 45) & (dados_forn['Cobertura Atual'] <= 70)])
        acima_71 = len(dados_forn[dados_forn['Cobertura Atual'] >= 71])
        
        perc_ate_44 = (ate_44 / total_itens) * 100
        perc_45_70 = (entre_45_70 / total_itens) * 100
        perc_acima_71 = (acima_71 / total_itens) * 100
        
        fornecedores_analise.append({
            'fornecedor': fornecedor,
            'total_itens': total_itens,
            'total_filiais': total_filiais,
            'valor_total': valor_total,
            'cobertura_media': cobertura_media,
            'ate_44': ate_44,
            'perc_ate_44': perc_ate_44,
            'entre_45_70': entre_45_70,
            'perc_45_70': perc_45_70,
            'acima_71': acima_71,
            'perc_acima_71': perc_acima_71
        })
    
    # Ordenar por % acima de 71 dias (mais críticos primeiro)
    fornecedores_analise.sort(key=lambda x: x['perc_acima_71'], reverse=True)
    
    # Preencher dados
    for row, forn_data in enumerate(fornecedores_analise, 4):
        dados = [
            forn_data['fornecedor'][:30],
            forn_data['total_itens'],
            forn_data['total_filiais'],
            f"R$ {forn_data['valor_total']:,.2f}",
            f"{forn_data['cobertura_media']:.1f}",
            forn_data['ate_44'],
            f"{forn_data['perc_ate_44']:.1f}%",
            forn_data['entre_45_70'],
            f"{forn_data['perc_45_70']:.1f}%",
            forn_data['acima_71'],
            f"{forn_data['perc_acima_71']:.1f}%"
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            
            # Colorir baseado na criticidade
            if forn_data['perc_acima_71'] > 50:
                cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
            elif forn_data['perc_acima_71'] > 25:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adicionar detalhamento por filial dentro de cada fornecedor crítico
    row_atual = len(fornecedores_analise) + 6
    
    # Focar apenas nos fornecedores mais críticos (top 10)
    fornecedores_criticos = [f for f in fornecedores_analise if f['perc_acima_71'] > 25][:10]
    
    if fornecedores_criticos:
        ws.cell(row=row_atual, column=1, value="DETALHAMENTO DOS FORNECEDORES MAIS CRÍTICOS").font = Font(bold=True, size=14)
        ws.cell(row=row_atual, column=1).fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
        ws.merge_cells(f'A{row_atual}:K{row_atual}')
        row_atual += 2
        
        for forn_data in fornecedores_criticos:
            fornecedor = forn_data['fornecedor']
            dados_forn = df[df['Fornecedor'] == fornecedor]
            
            # Título do fornecedor
            ws.cell(row=row_atual, column=1, value=f"FORNECEDOR: {fornecedor}").font = Font(bold=True, size=12)
            ws.cell(row=row_atual, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws.merge_cells(f'A{row_atual}:K{row_atual}')
            
            row_atual += 2
            
            # Cabeçalhos do detalhamento
            headers_det = ['Filial', 'Itens', 'Valor (R$)', 'Até 44d', '% 44d', '45-70d', '% 45-70d', 'Acima 71d', '% 71d', 'Cobertura Média']
            for col, header in enumerate(headers_det, 1):
                cell = ws.cell(row=row_atual, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            row_atual += 1
            
            # Analisar filiais do fornecedor
            for filial in dados_forn['Filial'].unique():
                dados_forn_filial = dados_forn[dados_forn['Filial'] == filial]
                total_itens_filial = len(dados_forn_filial)
                valor_filial = dados_forn_filial['Saldo Pedido'].sum()
                cobertura_media_filial = dados_forn_filial['Cobertura Atual'].mean()
                
                ate_44_filial = len(dados_forn_filial[dados_forn_filial['Cobertura Atual'] <= 44])
                entre_45_70_filial = len(dados_forn_filial[(dados_forn_filial['Cobertura Atual'] >= 45) & (dados_forn_filial['Cobertura Atual'] <= 70)])
                acima_71_filial = len(dados_forn_filial[dados_forn_filial['Cobertura Atual'] >= 71])
                
                perc_ate_44_filial = (ate_44_filial / total_itens_filial) * 100
                perc_45_70_filial = (entre_45_70_filial / total_itens_filial) * 100
                perc_acima_71_filial = (acima_71_filial / total_itens_filial) * 100
                
                dados_filial = [
                    filial,
                    total_itens_filial,
                    f"R$ {valor_filial:,.2f}",
                    ate_44_filial,
                    f"{perc_ate_44_filial:.1f}%",
                    entre_45_70_filial,
                    f"{perc_45_70_filial:.1f}%",
                    acima_71_filial,
                    f"{perc_acima_71_filial:.1f}%",
                    f"{cobertura_media_filial:.1f}"
                ]
                
                for col, valor in enumerate(dados_filial, 1):
                    cell = ws.cell(row=row_atual, column=col, value=valor)
                    
                    # Colorir baseado na criticidade da filial
                    if perc_acima_71_filial > 50:
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif perc_acima_71_filial > 25:
                        cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
                
                row_atual += 1
            
            row_atual += 2
    
    # Ajustar larguras
    larguras = [30, 10, 8, 15, 12, 10, 10, 12, 10, 10, 15]
    for col, largura in enumerate(larguras, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = largura
